from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


OUTPUT = Path(__file__).parents[1] / "frontend" / "sample-bulk.xlsx"


def build_sample(output: Path = OUTPUT) -> Path:
    headers = [
        "Product", "Entity", "Campaign ID", "Ad Group ID", "Portfolio ID",
        "Campaign Name (Informational only)", "Ad Group Name (Informational only)", "Portfolio Name (Informational only)",
        "SKU", "ASIN (Informational only)", "Match Type", "Placement",
        "Product Targeting Expression", "Keyword Text", "Impressions", "Clicks", "Spend", "Sales", "Orders",
    ]
    rows = [
        [
            "Sponsored Products", "Keyword", "DEMO-1001", "DEMO-GROUP-1", "DEMO-PORT-1",
            "Demo Carry-on Phrase", "Demo Group A", "Demo Portfolio", "DEMO-SKU-A", "B0DEMO0001",
            "Phrase", None, None, "carry on luggage", 12500, 315, 264.4, 1189.5, 17,
        ],
        [
            "Sponsored Products", "Keyword", "DEMO-1001", "DEMO-GROUP-1", "DEMO-PORT-1",
            "Demo Carry-on Phrase", "Demo Group A", "Demo Portfolio", "DEMO-SKU-A", "B0DEMO0001",
            "Phrase", None, None, "lightweight suitcase", 6800, 142, 113.7, 424.0, 6,
        ],
        [
            "Sponsored Products", "Keyword", "DEMO-1001", "DEMO-GROUP-1", "DEMO-PORT-1",
            "Demo Carry-on Phrase", "Demo Group A", "Demo Portfolio", "DEMO-SKU-A", "B0DEMO0001",
            "Exact", None, None, "luggage with wheels", 3100, 61, 58.2, 0, 0,
        ],
        [
            "Sponsored Products", "Bidding Adjustment", "DEMO-1001", None, "DEMO-PORT-1",
            "Demo Carry-on Phrase", None, "Demo Portfolio", None, None, None, "Placement Top",
            None, None, 8900, 228, 205.6, 1012.0, 15,
        ],
        [
            "Sponsored Products", "Bidding Adjustment", "DEMO-1001", None, "DEMO-PORT-1",
            "Demo Carry-on Phrase", None, "Demo Portfolio", None, None, None, "Placement Product Page",
            None, None, 5200, 96, 81.3, 302.0, 4,
        ],
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sponsored Products Campaigns"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:S{len(rows) + 1}"

    search = workbook.create_sheet("SP Search Term Report")
    search.append([
        "Product", "Campaign ID", "Ad Group ID", "Campaign Name (Informational only)",
        "Ad Group Name (Informational only)", "Customer Search Term", "Impressions", "Clicks",
        "Spend", "Sales", "Orders",
    ])
    search.append(["Sponsored Products", "DEMO-1001", "DEMO-GROUP-1", "Demo Carry-on Phrase", "Demo Group A", "carry on luggage", 9000, 250, 210, 900, 13])
    search.append(["Sponsored Products", "DEMO-1001", "DEMO-GROUP-1", "Demo Carry-on Phrase", "Demo Group A", "lightweight suitcase", 5000, 100, 80, 300, 4])
    search.append(["Sponsored Products", "DEMO-1001", "DEMO-GROUP-1", "Demo Carry-on Phrase", "Demo Group A", "luggage with wheels", 2000, 40, 35, 0, 0])

    portfolios = workbook.create_sheet("Portfolios")
    portfolios.append([
        "Product", "Entity", "Portfolio ID", "Portfolio Name", "Budget Amount", "Budget Currency Code",
        "Budget Policy", "Budget Start Date", "Budget End Date", "State (Informational only)",
    ])
    portfolios.append(["Portfolios", "Portfolio", "DEMO-PORT-1", "Demo Portfolio", 150, "USD", "dateRange", "20260601", "20260630", "enabled"])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


if __name__ == "__main__":
    print(build_sample())
