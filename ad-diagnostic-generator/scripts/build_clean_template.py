from __future__ import annotations

import argparse
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def remove_chart_external_data(source: Path, output: Path) -> None:
    """Remove WPS externalData nodes that openpyxl cannot parse safely."""
    with zipfile.ZipFile(source, "r") as source_zip, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename.startswith("xl/charts/chart") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for parent in root.iter():
                    for child in list(parent):
                        if child.tag == f"{{{CHART_NS}}}externalData":
                            parent.remove(child)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target_zip.writestr(item, data)


def clear_range(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None


def build_clean_template(source: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temp_dir:
        sanitized = Path(temp_dir) / "sanitized.xlsx"
        remove_chart_external_data(source, sanitized)
        workbook = load_workbook(sanitized, data_only=False, read_only=False, keep_links=False)

        budget = workbook["广告组合预算"]
        clear_range(budget, 4, 19, 1, 32)
        budget.freeze_panes = "C4"
        budget.auto_filter.ref = "A3:AF4"
        budget["AF3"] = 0.8
        budget["B20"] = "天数"
        budget["C20"] = 30
        budget["B21"] = "日花费"
        budget["C21"] = "=IFERROR(L2/C20,0)"

        bulk = workbook["Bulk"]
        clear_range(bulk, 15, bulk.max_row, 1, 57)
        if bulk.max_row > 15:
            bulk.delete_rows(16, bulk.max_row - 15)
        bulk.freeze_panes = "K15"
        bulk.auto_filter.ref = "A14:BE15"

        bidding = workbook["BidingAdjustment"]
        clear_range(bidding, 6, bidding.max_row, 1, 54)
        if bidding.max_row > 6:
            bidding.delete_rows(7, bidding.max_row - 6)
        bidding.freeze_panes = "N6"
        bidding.auto_filter.ref = "A5:BB6"

        visual = workbook["广告位可视化"]
        visual.freeze_panes = "B2"
        visual.auto_filter.ref = "B1:Y20"

        search = workbook["搜索词模板"]
        clear_range(search, 4, search.max_row, 1, 43)
        if search.max_row > 4:
            search.delete_rows(5, search.max_row - 4)
        search.freeze_panes = "B4"
        search.auto_filter.ref = "A3:AQ4"

        for sheet_name in ("词频分析(全部搜索词）", "词频分析 (不出单搜素词)"):
            sheet = workbook[sheet_name]
            clear_range(sheet, 2, sheet.max_row, 1, 19)
            if sheet.max_row > 2:
                sheet.delete_rows(3, sheet.max_row - 2)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = "A1:M2"

        asin = workbook["LX_Asin"]
        clear_range(asin, 2, asin.max_row, 1, 2)
        if asin.max_row > 2:
            asin.delete_rows(3, asin.max_row - 2)
        asin.freeze_panes = "A2"
        asin.auto_filter.ref = "A1:B2"

        niche_word = workbook["NicheWord"]
        niche_word["A1"] = "细分市场名称：未提供（需额外上传 ABA / Opportunity Explorer 数据）"
        niche_word["A3"] = None
        clear_range(niche_word, 6, niche_word.max_row, 1, 22)
        clear_range(niche_word, 1, 4, 13, 16)
        niche_word["M1"] = "细分市场总点击数量"
        niche_word["N1"] = "N/A"
        if niche_word.max_row > 6:
            niche_word.delete_rows(7, niche_word.max_row - 6)
        niche_word.freeze_panes = "A6"
        niche_word.auto_filter.ref = "A5:T6"

        niche_asin = workbook["NicheAsin"]
        niche_asin["A1"] = "细分市场名称：未提供（需额外上传 ABA / Opportunity Explorer 数据）"
        niche_asin["A3"] = None
        clear_range(niche_asin, 6, niche_asin.max_row, 1, 16)
        if niche_asin.max_row > 6:
            niche_asin.delete_rows(7, niche_asin.max_row - 6)
        niche_asin.freeze_panes = "A6"
        niche_asin.auto_filter.ref = "A5:L6"

        negative = workbook["否词库"]
        clear_range(negative, 2, negative.max_row, 1, 5)
        negative.freeze_panes = "A2"

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a de-identified template for the ad diagnostic app.")
    parser.add_argument("source", type=Path, help="Path to the private source template")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parents[1] / "assets" / "广告诊断表_空白模板.xlsx",
    )
    args = parser.parse_args()
    build_clean_template(args.source, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
