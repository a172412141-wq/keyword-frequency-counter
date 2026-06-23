from __future__ import annotations

import io
import math
import re
import unicodedata
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date
from itertools import islice
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TEMPLATE_PATH = Path(__file__).parent / "assets" / "广告诊断表_空白模板.xlsx"

PARENT_FIELDS = (
    "ASIN (Informational only)",
    "SKU",
    "Campaign Name",
    "Campaign Name (Informational only)",
    "Ad Group Name",
    "Ad Group Name (Informational only)",
)

METRIC_FIELDS = ("Impressions", "Clicks", "Spend", "Sales", "Orders")
SEARCH_TERM_FIELDS = (
    "Customer Search Term",
    "Keyword Text",
    "Product Targeting Expression",
)

HEADER_ALIASES = {
    "7 day total sales": "Sales",
    "7 day total sales (usd)": "Sales",
    "14 day total sales": "Sales",
    "sales (usd)": "Sales",
    "7 day total orders (#)": "Orders",
    "14 day total orders (#)": "Orders",
    "orders (#)": "Orders",
    "cost": "Spend",
    "spend (usd)": "Spend",
    "click thru rate": "Click-through Rate",
    "click through rate": "Click-through Rate",
    "conversion rate": "Conversion Rate",
    "advertised sku": "SKU",
    "advertised asin": "ASIN (Informational only)",
    "asin": "ASIN (Informational only)",
    "portfolio name": "Portfolio Name (Informational only)",
    "campaign": "Campaign Name",
    "ad group": "Ad Group Name",
    "customer search term": "Customer Search Term",
}


class InputValidationError(ValueError):
    """Raised when an uploaded workbook cannot safely produce a report."""


@dataclass(frozen=True)
class BulkInput:
    sheet_name: str
    header_row: int
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]
    search_term_sheet_name: str | None = None
    search_term_headers: tuple[str, ...] = ()
    search_term_rows: tuple[dict[str, Any], ...] = ()
    portfolio_sheet_name: str | None = None
    portfolio_headers: tuple[str, ...] = ()
    portfolio_rows: tuple[dict[str, Any], ...] = ()


@dataclass(frozen=True)
class GenerationStats:
    selected_field: str
    selected_values: tuple[str, ...]
    input_rows: int
    filtered_rows: int
    bidding_rows: int
    search_terms: int
    portfolios: int


REMOVED_OUTPUT_SHEETS = ("LX_Asin", "NicheWord", "NicheAsin", "H1F广告架构")


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = re.sub(r"\s+", " ", text)
    key = text.casefold()
    return HEADER_ALIASES.get(key, text)


def _header_key(value: Any) -> str:
    return normalize_header(value).casefold()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _as_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return 0.0
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return 0.0
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _read_source(source: bytes | bytearray | io.BytesIO | str | Path):
    if isinstance(source, (str, Path)):
        return load_workbook(source, data_only=False, read_only=True, keep_links=False)
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    return load_workbook(source, data_only=False, read_only=True, keep_links=False)


def _sheet_profiles(workbook) -> list[tuple[Any, int, list[str], set[str]]]:
    profiles = []
    for sheet in workbook.worksheets:
        # Amazon's generated Bulk files can incorrectly declare A1:A1 as the
        # worksheet dimension. Resetting it makes read-only streaming inspect
        # the actual XML rows instead of silently returning one cell.
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        for row_index, values in enumerate(islice(sheet.iter_rows(values_only=True), 30), 1):
            headers = [normalize_header(value) for value in values]
            keys = {_header_key(header) for header in headers if header}
            if keys:
                profiles.append((sheet, row_index, headers, keys))
    return profiles


def _records_from_profile(profile: tuple[Any, int, list[str], set[str]]) -> list[dict[str, Any]]:
    sheet, header_row, headers, _ = profile
    records: list[dict[str, Any]] = []
    blank_streak = 0
    for row_index, values in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_index <= header_row:
            continue
        values = tuple(values[:len(headers)])
        if all(_is_blank(value) for value in values):
            blank_streak += 1
            if blank_streak >= 50:
                break
            continue
        blank_streak = 0
        records.append({header: value for header, value in zip(headers, values) if header})
    return records


def _best_profile(
    profiles: Sequence[tuple[Any, int, list[str], set[str]]],
    required: set[str],
    preferred_sheet: str,
) -> tuple[Any, int, list[str], set[str]] | None:
    candidates = [profile for profile in profiles if required.issubset(profile[3])]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda profile: (
            profile[0].title.casefold() == preferred_sheet.casefold(),
            len(profile[3]),
            -profile[1],
        ),
    )


def read_bulk_input(source: bytes | bytearray | io.BytesIO | str | Path) -> BulkInput:
    workbook = _read_source(source)
    profiles = _sheet_profiles(workbook)
    metric_keys = {_header_key(field) for field in METRIC_FIELDS}
    main = _best_profile(
        profiles,
        {"entity", "campaign id", *metric_keys},
        "Sponsored Products Campaigns",
    )
    # Backward compatibility for simplified single-sheet files without IDs.
    if main is None:
        main_candidates = [
            profile for profile in profiles
            if "entity" in profile[3] and metric_keys.issubset(profile[3])
        ]
        main = max(main_candidates, key=lambda profile: len(profile[3]), default=None)
    if main is None:
        incomplete_candidates = [profile for profile in profiles if "entity" in profile[3]]
        main = max(
            incomplete_candidates,
            key=lambda profile: (
                profile[0].title.casefold() == "sponsored products campaigns",
                len(metric_keys & profile[3]),
                len(profile[3]),
            ),
            default=None,
        )
    if main is None:
        raise InputValidationError("未识别到 Sponsored Products Campaigns。请上传 Amazon Ads 控制台导出的标准 Bulk .xlsx。")

    search = _best_profile(
        profiles,
        {"customer search term", "campaign id", *metric_keys},
        "SP Search Term Report",
    )
    portfolio = _best_profile(
        profiles,
        {"portfolio id", "portfolio name (informational only)", "budget policy"},
        "Portfolios",
    )

    sheet, header_row, headers, _ = main
    records = _records_from_profile(main)
    search_records = _records_from_profile(search) if search else []
    portfolio_records = _records_from_profile(portfolio) if portfolio else []

    if not records:
        raise InputValidationError("Bulk 表只有表头，没有可用数据行。")

    available = {_header_key(header) for header in headers if header}
    missing_metrics = [field for field in METRIC_FIELDS if _header_key(field) not in available]
    if "entity" not in available:
        raise InputValidationError("缺少必要字段：Entity。")
    if missing_metrics:
        raise InputValidationError("缺少必要指标字段：" + "、".join(missing_metrics))
    if not any(_header_key(field) in available for field in PARENT_FIELDS):
        raise InputValidationError("缺少父体筛选字段：ASIN、SKU、Campaign Name 或 Ad Group Name。")
    if not any(_as_text(row.get(field)) for row in records for field in PARENT_FIELDS):
        raise InputValidationError("父体筛选字段存在，但没有可选择的 ASIN、SKU、Campaign 或 Ad Group 值。")
    if not search_records and not any(_header_key(field) in available for field in SEARCH_TERM_FIELDS):
        raise InputValidationError("缺少搜索词来源字段：Customer Search Term、Keyword Text 或 Product Targeting Expression。")

    return BulkInput(
        sheet_name=sheet.title,
        header_row=header_row,
        headers=tuple(header for header in headers if header),
        rows=tuple(records),
        search_term_sheet_name=search[0].title if search else None,
        search_term_headers=tuple(header for header in search[2] if header) if search else (),
        search_term_rows=tuple(search_records),
        portfolio_sheet_name=portfolio[0].title if portfolio else None,
        portfolio_headers=tuple(header for header in portfolio[2] if header) if portfolio else (),
        portfolio_rows=tuple(portfolio_records),
    )


def parent_options(bulk: BulkInput) -> dict[str, list[str]]:
    options: dict[str, list[str]] = {}
    for field in PARENT_FIELDS:
        values = sorted({_as_text(row.get(field)) for row in bulk.rows if _as_text(row.get(field))}, key=str.casefold)
        if values:
            options[field] = values
    return options


def _selection_values(values: str | Sequence[str]) -> tuple[str, ...]:
    source = (values,) if isinstance(values, str) else values
    result: list[str] = []
    seen: set[str] = set()
    for value in source:
        cleaned = _as_text(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            result.append(cleaned)
    return tuple(result)


def filter_parent_rows(
    rows: Sequence[Mapping[str, Any]],
    field: str,
    values: str | Sequence[str],
) -> list[dict[str, Any]]:
    wanted_values = tuple(value.casefold() for value in _selection_values(values))
    if not wanted_values:
        return []
    fuzzy = field.startswith("Campaign Name") or field.startswith("Ad Group Name")
    directly_matched: list[dict[str, Any]] = []
    for row in rows:
        actual = _as_text(row.get(field)).casefold()
        matched = any(wanted in actual for wanted in wanted_values) if fuzzy else actual in wanted_values
        if matched:
            directly_matched.append(dict(row))

    if not directly_matched:
        return directly_matched

    # Placement/bidding rows usually have no ASIN or SKU. Pull them back into
    # the selected product set through the campaign/ad-group identifiers found
    # on the directly matched rows.
    relation_fields = (
        "Campaign ID",
        "Campaign Name",
        "Campaign Name (Informational only)",
        "Ad Group ID",
        "Ad Group Name",
        "Ad Group Name (Informational only)",
    )
    related = {
        relation_field: {
            _as_text(row.get(relation_field)).casefold()
            for row in directly_matched
            if _as_text(row.get(relation_field))
        }
        for relation_field in relation_fields
    }
    strong_fields = [field for field in ("Campaign ID", "Ad Group ID") if related[field]]
    match_fields = strong_fields or [field for field in relation_fields[1:] if related[field]]
    selected: list[dict[str, Any]] = []
    for row in rows:
        actual = _as_text(row.get(field)).casefold()
        direct = any(wanted in actual for wanted in wanted_values) if fuzzy else actual in wanted_values
        shares_relation = any(
            _as_text(row.get(relation_field)).casefold() in values
            for relation_field in match_fields
            for values in (related[relation_field],)
            if values and _as_text(row.get(relation_field))
        )
        if direct or shares_relation:
            selected.append(dict(row))
    return selected


def filter_related_rows(
    rows: Sequence[Mapping[str, Any]],
    selected_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """Match auxiliary report rows to the selected campaign/ad-group set."""
    relation_fields = (
        "Campaign ID",
        "Ad Group ID",
        "Campaign Name (Informational only)",
        "Ad Group Name (Informational only)",
    )
    related = {
        field: {
            _as_text(row.get(field)).casefold()
            for row in selected_rows
            if _as_text(row.get(field))
        }
        for field in relation_fields
    }
    strong_fields = [field for field in ("Campaign ID", "Ad Group ID") if related[field]]
    match_fields = strong_fields or [field for field in relation_fields[2:] if related[field]]
    return [
        dict(row)
        for row in rows
        if any(
            _as_text(row.get(field)).casefold() in related[field]
            for field in match_fields
            if _as_text(row.get(field))
        )
    ]


def safe_filename(parent_value: str, today: date | None = None) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\x00-\x1f]+", "_", parent_value).strip(" ._")
    cleaned = re.sub(r"\s+", "_", cleaned)[:80] or "全部数据"
    return f"广告诊断表_{cleaned}_{(today or date.today()).isoformat()}.xlsx"


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    if target_row == source_row:
        return
    source_dim = sheet.row_dimensions[source_row]
    target_dim = sheet.row_dimensions[target_row]
    target_dim.height = source_dim.height
    target_dim.hidden = source_dim.hidden
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _clear_values(sheet, start_row: int, end_row: int, max_col: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=max(start_row, end_row), min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def _write_records(sheet, start_row: int, headers: Sequence[Any], rows: Sequence[Mapping[str, Any]], max_col: int) -> int:
    end_row = start_row + max(len(rows), 1) - 1
    _clear_values(sheet, start_row, max(sheet.max_row, end_row), max_col)
    for offset, row in enumerate(rows):
        target_row = start_row + offset
        _copy_row_style(sheet, start_row, target_row, max_col)
        for col, header in enumerate(headers, 1):
            if header:
                sheet.cell(target_row, col).value = row.get(normalize_header(header))
    return start_row + len(rows) - 1 if rows else start_row


def _excel_quote(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _write_bulk_sheet(workbook, rows: Sequence[Mapping[str, Any]]) -> int:
    sheet = workbook["Bulk"]
    headers = [sheet.cell(14, col).value for col in range(1, 58)]
    last_row = _write_records(sheet, 15, headers, rows, 57)
    sheet.freeze_panes = "K15"
    sheet.auto_filter.ref = f"A14:BE{last_row}"

    labels = ["Exact", "Phrase", "Broad", "Product Targeting", "asin=", "asin-expanded=", "category=", "complements", "loose-match", "close-match", "substitutes"]
    for index, label in enumerate(labels, 1):
        sheet.cell(index, 32).value = label

    metric_columns = {44: "AR", 45: "AS", 47: "AU", 48: "AV", 49: "AW"}
    for summary_row in range(1, 12):
        for col, metric_letter in metric_columns.items():
            if summary_row <= 3:
                formula = f'=SUMIF($AF$15:$AF${last_row},$AF{summary_row},${metric_letter}$15:${metric_letter}${last_row})'
            elif summary_row == 4:
                formula = f'=SUMIF($B$15:$B${last_row},$AF4,${metric_letter}$15:${metric_letter}${last_row})'
            else:
                formula = (
                    f'=SUMPRODUCT(--ISNUMBER(SEARCH($AF{summary_row},$AJ$15:$AJ${last_row})),'
                    f'${metric_letter}$15:${metric_letter}${last_row})'
                )
            sheet.cell(summary_row, col).value = formula
    return last_row


def _write_bidding_sheet(workbook, rows: Sequence[Mapping[str, Any]]) -> int:
    sheet = workbook["BidingAdjustment"]
    bidding = [row for row in rows if _as_text(row.get("Entity")).casefold() == "bidding adjustment"]
    headers = [sheet.cell(5, col).value for col in range(1, 55)]
    last_row = _write_records(sheet, 6, headers, bidding, 54)
    sheet.freeze_panes = "N6"
    sheet.auto_filter.ref = f"A5:BB{last_row}"

    labels = ["Placement Product Page", "Placement Rest Of Search", "Placement Top", "Placement Amazon Business"]
    metric_columns = {44: "AR", 45: "AS", 47: "AU", 48: "AV", 49: "AW"}
    for summary_row, label in enumerate(labels, 1):
        sheet.cell(summary_row, 34).value = label
        for col, metric_letter in metric_columns.items():
            sheet.cell(summary_row, col).value = (
                f'=SUMIF($AH$6:$AH${last_row},$AH{summary_row},${metric_letter}$6:${metric_letter}${last_row})'
            )
    return len(bidding)


def _write_visualization(workbook) -> None:
    sheet = workbook["广告位可视化"]
    source = _excel_quote("BidingAdjustment")
    for row in range(2, 6):
        source_row = row - 1
        for col, letter in zip(range(3, 8), ("AR", "AS", "AU", "AV", "AW")):
            sheet.cell(row, col).value = f"={source}!{letter}{source_row}"
        _write_metric_formulas(sheet, row)
    for col in range(3, 8):
        letter = get_column_letter(col)
        sheet.cell(6, col).value = f"=SUM({letter}2:{letter}5)"
    _write_metric_formulas(sheet, 6)

    bulk_rows = (1, 2, 3, 5, 6, 7, 8, 9, 10, 11)
    bulk = _excel_quote("Bulk")
    for row, source_row in zip(range(9, 19), bulk_rows):
        for col, letter in zip(range(3, 8), ("AR", "AS", "AU", "AV", "AW")):
            sheet.cell(row, col).value = f"={bulk}!{letter}{source_row}"
        _write_metric_formulas(sheet, row)
    for col in range(3, 8):
        letter = get_column_letter(col)
        sheet.cell(19, col).value = f"=SUM({letter}9:{letter}18)"
    _write_metric_formulas(sheet, 19)


def _write_metric_formulas(sheet, row: int) -> None:
    sheet.cell(row, 8).value = f"=IFERROR(E{row}/D{row},0)"
    sheet.cell(row, 9).value = f"=IFERROR(D{row}/C{row},0)"
    sheet.cell(row, 10).value = f"=IFERROR(G{row}/D{row},0)"
    sheet.cell(row, 11).value = f"=IFERROR(E{row}/F{row},0)"
    sheet.cell(row, 12).value = f"=IFERROR(E{row}/C{row}*1000,0)"
    sheet.cell(row, 13).value = f"=IFERROR(E{row}/G{row},0)"
    sheet.cell(row, 14).value = f"=IFERROR(F{row}/G{row},0)"
    total_row = 6 if row <= 6 else 19
    sheet.cell(row, 15).value = f"=IFERROR(F{row}/$F${total_row},0)"
    sheet.cell(row, 16).value = f"=IFERROR(E{row}/$E${total_row},0)"
    sheet.cell(row, 17).value = f"=O{row}-P{row}"
    sheet.cell(row, 18).value = f"=K{row}"
    sheet.cell(row, 19).value = f"=IFERROR(1/K{row},0)"


def _search_term_for(row: Mapping[str, Any]) -> str:
    for field in SEARCH_TERM_FIELDS:
        value = _as_text(row.get(field))
        if value:
            return value
    return ""


def _aggregate_search_terms(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        term = _search_term_for(row)
        if not term:
            continue
        key = term.casefold()
        item = grouped.setdefault(
            key,
            {"term": term, "Impressions": 0.0, "Clicks": 0.0, "Spend": 0.0, "Sales": 0.0, "Orders": 0.0},
        )
        for metric in METRIC_FIELDS:
            item[metric] += _as_number(row.get(metric))
    return sorted(grouped.values(), key=lambda item: (-item["Sales"], -item["Spend"], item["term"].casefold()))


def _write_search_terms(workbook, aggregated: Sequence[Mapping[str, Any]]) -> int:
    sheet = workbook["搜索词模板"]
    max_col = 43
    last_row = 3 + max(len(aggregated), 1)
    _clear_values(sheet, 4, max(sheet.max_row, last_row), max_col)
    for offset, item in enumerate(aggregated):
        row = 4 + offset
        _copy_row_style(sheet, 4, row, max_col)
        sheet.cell(row, 1).value = item["term"]
        for col, metric in zip(range(2, 7), METRIC_FIELDS):
            sheet.cell(row, col).value = item[metric]
        formulas = {
            7: f"=IFERROR(D{row}/C{row},0)",
            8: f"=IFERROR(C{row}/B{row},0)",
            9: f"=IFERROR(F{row}/C{row},0)",
            10: f"=IFERROR(D{row}/E{row},0)",
            11: f"=IFERROR(D{row}/B{row}*1000,0)",
            12: f"=IFERROR(D{row}/F{row},0)",
            13: f"=IFERROR(E{row}/F{row},0)",
            14: f'=IF(LEFT(UPPER(A{row}),2)="B0","ASIN","关键词")',
            15: f"=IFERROR(F{row}/$F$1,0)",
            16: f"=IFERROR(E{row}/$E$1,0)",
            17: f"=IFERROR(D{row}/$D$1,0)",
            18: f"=P{row}-Q{row}",
        }
        for col, formula in formulas.items():
            sheet.cell(row, col).value = formula

    data_end = 3 + len(aggregated) if aggregated else 4
    for col in range(2, 7):
        letter = get_column_letter(col)
        sheet.cell(1, col).value = f"=SUM({letter}4:{letter}{data_end})"
        sheet.cell(2, col).value = f"=SUBTOTAL(9,{letter}4:{letter}{data_end})"
    top_formulas = {
        7: "=IFERROR(D1/C1,0)", 8: "=IFERROR(C1/B1,0)", 9: "=IFERROR(F1/C1,0)",
        10: "=IFERROR(D1/E1,0)", 11: "=IFERROR(D1/B1*1000,0)", 12: "=IFERROR(D1/F1,0)",
        13: "=IFERROR(E1/F1,0)", 15: "=IFERROR(F1/F1,0)", 16: "=IFERROR(E1/E1,0)",
        17: "=IFERROR(D1/D1,0)", 18: "=P1-Q1",
    }
    filter_formulas = {
        7: "=IFERROR(D2/C2,0)", 8: "=IFERROR(C2/B2,0)", 9: "=IFERROR(F2/C2,0)",
        10: "=IFERROR(D2/E2,0)", 11: "=IFERROR(D2/B2*1000,0)", 12: "=IFERROR(D2/F2,0)",
        13: "=IFERROR(E2/F2,0)", 14: "=IFERROR(L2/L1,0)", 15: "=IFERROR(F2/F1,0)",
        16: "=IFERROR(E2/E1,0)", 17: "=IFERROR(D2/D1,0)", 18: "=P2-Q2",
    }
    for col, formula in top_formulas.items():
        sheet.cell(1, col).value = formula
    for col, formula in filter_formulas.items():
        sheet.cell(2, col).value = formula
    sheet.freeze_panes = "B4"
    sheet.auto_filter.ref = f"A3:AQ{data_end}"
    return len(aggregated)


def _tokens(text: str) -> list[str]:
    normalized = unicodedata.normalize("NFKC", text).casefold().replace("’", "'")
    return re.findall(r"[a-z0-9]+(?:['-][a-z0-9]+)*|[\u3400-\u9fff]+", normalized)


def _word_counts(aggregated: Sequence[Mapping[str, Any]], no_orders_only: bool) -> Counter[str]:
    counts: Counter[str] = Counter()
    for item in aggregated:
        if no_orders_only and _as_number(item["Orders"]) != 0:
            continue
        counts.update(_tokens(_as_text(item["term"])))
    return counts


def _write_word_frequency(workbook, aggregated: Sequence[Mapping[str, Any]], sheet_name: str, no_orders_only: bool) -> int:
    sheet = workbook[sheet_name]
    counts = _word_counts(aggregated, no_orders_only)
    items = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    total = sum(counts.values()) or 1
    last_row = 1 + max(len(items), 1)
    _clear_values(sheet, 2, max(sheet.max_row, last_row), 19)
    search_end = 3 + len(aggregated) if aggregated else 4
    search = _excel_quote("搜索词模板")
    for offset, (token, count) in enumerate(items):
        row = 2 + offset
        _copy_row_style(sheet, 2, row, 19)
        sheet.cell(row, 1).value = token
        sheet.cell(row, 2).value = count
        sheet.cell(row, 3).value = count / total
        sheet.cell(row, 3).number_format = "0.00%"
        for col, source_col in zip(range(4, 9), "BCDEF"):
            if no_orders_only:
                formula = (
                    f'=SUMPRODUCT(--ISNUMBER(SEARCH($A{row},{search}!$A$4:$A${search_end})),'
                    f'--({search}!$F$4:$F${search_end}=0),{search}!${source_col}$4:${source_col}${search_end})'
                )
            else:
                formula = (
                    f'=SUMPRODUCT(--ISNUMBER(SEARCH($A{row},{search}!$A$4:$A${search_end})),'
                    f'{search}!${source_col}$4:${source_col}${search_end})'
                )
            sheet.cell(row, col).value = formula
        sheet.cell(row, 9).value = f"=IFERROR(F{row}/E{row},0)"
        sheet.cell(row, 10).value = f"=IFERROR(E{row}/D{row},0)"
        sheet.cell(row, 11).value = f"=IFERROR(H{row}/E{row},0)"
        sheet.cell(row, 12).value = f"=IFERROR(F{row}/G{row},0)"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{last_row}"
    return len(items)


def _portfolio_name(row: Mapping[str, Any]) -> str:
    return _as_text(row.get("Portfolio Name (Informational only)")) or "未分配广告组合"


def _aggregate_portfolios(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    campaign_rows = [row for row in rows if _as_text(row.get("Entity")).casefold() == "campaign"]
    source_rows = campaign_rows or [
        row for row in rows if _as_text(row.get("Entity")).casefold() != "bidding adjustment"
    ]
    grouped: dict[str, dict[str, Any]] = {}
    campaigns: dict[str, set[str]] = defaultdict(set)
    portfolio_ids: dict[str, set[str]] = defaultdict(set)
    for row in source_rows:
        name = _portfolio_name(row)
        item = grouped.setdefault(name, {metric: 0.0 for metric in METRIC_FIELDS})
        for metric in METRIC_FIELDS:
            item[metric] += _as_number(row.get(metric))
        campaign = _as_text(row.get("Campaign Name (Informational only)")) or _as_text(row.get("Campaign Name"))
        if campaign:
            campaigns[name].add(campaign)
        portfolio_id = _as_text(row.get("Portfolio ID"))
        if portfolio_id:
            portfolio_ids[name].add(portfolio_id)
    result = []
    for name, metrics in grouped.items():
        result.append({
            "name": name,
            "portfolio_ids": tuple(sorted(portfolio_ids[name])),
            "campaigns": len(campaigns[name]),
            **metrics,
        })
    return sorted(result, key=lambda item: (-item["Spend"], item["name"].casefold()))


def _write_portfolios(
    workbook,
    rows: Sequence[Mapping[str, Any]],
    portfolio_rows: Sequence[Mapping[str, Any]] = (),
) -> int:
    sheet = workbook["广告组合预算"]
    items = _aggregate_portfolios(rows)
    metadata_by_id = {
        _as_text(row.get("Portfolio ID")).casefold(): row
        for row in portfolio_rows
        if _as_text(row.get("Portfolio ID"))
    }
    metadata_by_name = {
        _portfolio_name(row).casefold(): row
        for row in portfolio_rows
        if _portfolio_name(row) != "未分配广告组合"
    }
    last_row = 3 + max(len(items), 1)
    _clear_values(sheet, 4, max(19, last_row), 32)
    for offset, item in enumerate(items):
        row = 4 + offset
        _copy_row_style(sheet, 4, row, 32)
        metadata = next(
            (metadata_by_id[portfolio_id.casefold()] for portfolio_id in item["portfolio_ids"] if portfolio_id.casefold() in metadata_by_id),
            metadata_by_name.get(item["name"].casefold(), {}),
        )
        sheet.cell(row, 2).value = item["name"]
        sheet.cell(row, 3).value = _as_text(metadata.get("State (Informational only)")) or "N/A"
        sheet.cell(row, 4).value = _as_text(metadata.get("Budget Policy")) or "N/A"
        sheet.cell(row, 5).value = metadata.get("Budget Amount") if not _is_blank(metadata.get("Budget Amount")) else "N/A"
        sheet.cell(row, 6).value = metadata.get("Budget Start Date") if not _is_blank(metadata.get("Budget Start Date")) else "N/A"
        sheet.cell(row, 7).value = metadata.get("Budget End Date") if not _is_blank(metadata.get("Budget End Date")) else "N/A"
        sheet.cell(row, 8).value = item["campaigns"]
        sheet.cell(row, 9).value = item["Impressions"]
        sheet.cell(row, 10).value = item["Clicks"]
        sheet.cell(row, 12).value = item["Spend"]
        sheet.cell(row, 14).value = item["Orders"]
        sheet.cell(row, 15).value = item["Sales"]
        formulas = {
            11: f"=IFERROR(J{row}/I{row},0)", 13: f"=IFERROR(L{row}/J{row},0)",
            16: f"=IFERROR(L{row}/O{row},0)", 17: f"=IFERROR(1/P{row},0)",
            24: f"=IFERROR(N{row}/J{row},0)", 25: f"=IFERROR(O{row}/N{row},0)",
            26: f"=IFERROR(O{row}/$O$2,0)", 27: f"=IFERROR(L{row}/$L$2,0)",
            28: f"=Z{row}-AA{row}", 29: f"=IFERROR(L{row}/30,0)",
            30: f"=AC{row}*30", 31: f"=IFERROR(O{row}/N{row},0)", 32: f"=AC{row}*$AF$3",
        }
        for col, formula in formulas.items():
            sheet.cell(row, col).value = formula

    data_end = 3 + len(items) if items else 4
    for summary_row, function in ((1, "SUM"), (2, "SUBTOTAL")):
        for col in (9, 10, 12, 14, 15):
            letter = get_column_letter(col)
            if function == "SUM":
                sheet.cell(summary_row, col).value = f"=SUM({letter}4:{letter}{data_end})"
            else:
                sheet.cell(summary_row, col).value = f"=SUBTOTAL(9,{letter}4:{letter}{data_end})"
        formulas = {
            11: f"=IFERROR(J{summary_row}/I{summary_row},0)",
            13: f"=IFERROR(L{summary_row}/J{summary_row},0)",
            16: f"=IFERROR(L{summary_row}/O{summary_row},0)",
            17: f"=IFERROR(1/P{summary_row},0)",
        }
        for col, formula in formulas.items():
            sheet.cell(summary_row, col).value = formula
    sheet["AF3"] = 0.8
    sheet.freeze_panes = "C4"
    sheet.auto_filter.ref = f"A3:AF{data_end}"
    return len(items)


def generate_diagnostic_workbook(
    bulk: BulkInput,
    selected_field: str,
    selected_values: str | Sequence[str],
    template_path: str | Path = TEMPLATE_PATH,
) -> tuple[bytes, GenerationStats]:
    if selected_field not in PARENT_FIELDS:
        raise InputValidationError(f"不支持的父体筛选字段：{selected_field}")
    normalized_values = _selection_values(selected_values)
    if not normalized_values:
        raise InputValidationError("请至少选择一个父体或产品组。")
    filtered = filter_parent_rows(bulk.rows, selected_field, normalized_values)
    if not filtered:
        raise InputValidationError("所选父体没有匹配到任何数据行，请更换筛选值。")

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"内置模板不存在：{template}")
    workbook = load_workbook(template, data_only=False, read_only=False, keep_links=False)

    _write_bulk_sheet(workbook, filtered)
    bidding_count = _write_bidding_sheet(workbook, filtered)
    _write_visualization(workbook)
    search_rows = filter_related_rows(bulk.search_term_rows, filtered) if bulk.search_term_rows else filtered
    aggregated = _aggregate_search_terms(search_rows)
    search_term_count = _write_search_terms(workbook, aggregated)
    _write_word_frequency(workbook, aggregated, "词频分析(全部搜索词）", False)
    _write_word_frequency(workbook, aggregated, "词频分析 (不出单搜素词)", True)
    portfolio_count = _write_portfolios(workbook, filtered, bulk.portfolio_rows)
    for sheet_name in REMOVED_OUTPUT_SHEETS:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), GenerationStats(
        selected_field=selected_field,
        selected_values=normalized_values,
        input_rows=len(bulk.rows),
        filtered_rows=len(filtered),
        bidding_rows=bidding_count,
        search_terms=search_term_count,
        portfolios=portfolio_count,
    )
