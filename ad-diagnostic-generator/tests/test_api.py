from __future__ import annotations

import io
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import app


client = TestClient(app)


def sample_bulk() -> bytes:
    return (Path(__file__).parents[1] / "frontend" / "sample-bulk.xlsx").read_bytes()


def test_frontend_and_health_are_served():
    response = client.get("/")
    assert response.status_code == 200
    assert "把一张 Bulk" in response.text
    assert client.get("/api/health").json() == {"status": "ok"}
    assert client.get("/static/styles.css").status_code == 200


def test_analyze_returns_parent_choices():
    response = client.post(
        "/api/analyze",
        files={"file": ("bulk.xlsx", sample_bulk(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["data_rows"] == 7
    assert body["summary"]["search_term_sheet_name"] == "SP Search Term Report"
    assert body["summary"]["search_term_rows"] == 4
    assert body["summary"]["portfolio_rows"] == 1
    asin = next(field for field in body["fields"] if field["name"] == "ASIN (Informational only)")
    assert asin["values"] == ["B0DEMO0001", "B0DEMO0002"]


def test_generate_combines_multiple_skus():
    response = client.post(
        "/api/generate",
        files={"file": ("bulk.xlsx", sample_bulk(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "selected_field": "SKU",
            "selected_values": ["DEMO-SKU-A", "DEMO-SKU-B"],
        },
    )
    assert response.status_code == 200
    stats = response.headers["x-generation-stats"]
    assert '"selected_values":["DEMO-SKU-A","DEMO-SKU-B"]' in stats
    workbook = load_workbook(io.BytesIO(response.content), data_only=False, keep_links=False)
    assert workbook["Bulk"]["W20"].value == "B0DEMO0002"
    assert workbook["搜索词模板"]["A4"].value in {"carry on luggage", "checked luggage"}


def test_generate_downloads_workbook():
    response = client.post(
        "/api/generate",
        files={"file": ("bulk.xlsx", sample_bulk(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"selected_field": "ASIN (Informational only)", "selected_values": "B0DEMO0001"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(response.content), data_only=False, keep_links=False)
    assert len(workbook.sheetnames) == 8
    assert not {"LX_Asin", "NicheWord", "NicheAsin", "H1F广告架构"} & set(workbook.sheetnames)
    assert workbook["Bulk"]["W15"].value == "B0DEMO0001"
    assert workbook["BidingAdjustment"]["AH6"].value == "Placement Top"
    assert workbook["搜索词模板"]["A4"].value == "carry on luggage"
    assert workbook["搜索词模板"]["B4"].value == 9000
    assert workbook["广告组合预算"]["C4"].value == "enabled"
    assert workbook["广告组合预算"]["D4"].value == "dateRange"
    assert workbook["广告组合预算"]["E4"].value == 150


def test_analyze_rejects_non_xlsx():
    response = client.post(
        "/api/analyze",
        files={"file": ("bulk.csv", b"a,b\n1,2", "text/csv")},
    )
    assert response.status_code == 415
