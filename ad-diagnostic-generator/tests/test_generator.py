from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from generator import (
    InputValidationError,
    TEMPLATE_PATH,
    generate_diagnostic_workbook,
    parent_options,
    read_bulk_input,
    safe_filename,
)


HEADERS = [
    "Product",
    "Entity",
    "Campaign Name (Informational only)",
    "Ad Group Name (Informational only)",
    "Portfolio Name (Informational only)",
    "SKU",
    "ASIN (Informational only)",
    "Match Type",
    "Placement",
    "Product Targeting Expression",
    "Keyword Text",
    "Customer Search Term",
    "Impressions",
    "Clicks",
    "Spend",
    "Sales",
    "Orders",
]


def workbook_bytes(rows: list[list[object]], headers: list[str] | None = None, header_row: int = 1) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sponsored Products Campaigns"
    for _ in range(1, header_row):
        sheet.append([])
    sheet.append(headers or HEADERS)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture()
def sample_rows() -> list[list[object]]:
    return [
        [
            "Sponsored Products", "Keyword", "Parent-A Phrase", "Group 1", "Portfolio A", "SKU-A",
            "B0PARENTA1", "Phrase", None, None, "carry on luggage", "carry on luggage", 1000, 50, 40, 200, 5,
        ],
        [
            "Sponsored Products", "Bidding Adjustment", "Parent-A Phrase", "Group 1", "Portfolio A", None,
            None, None, "Placement Top", None, None, None, 500, 25, 20, 120, 3,
        ],
        [
            "Sponsored Products", "Keyword", "Other Campaign", "Group 2", "Portfolio B", "SKU-B",
            "B0OTHER001", "Exact", None, None, "laptop bag", "laptop bag", 900, 30, 35, 100, 2,
        ],
    ]


def test_reads_header_row_and_parent_options(sample_rows):
    bulk = read_bulk_input(workbook_bytes(sample_rows, header_row=14))
    assert bulk.header_row == 14
    assert len(bulk.rows) == 3
    options = parent_options(bulk)
    assert options["ASIN (Informational only)"] == ["B0OTHER001", "B0PARENTA1"]
    assert "SKU-A" in options["SKU"]


def test_generates_filtered_template_workbook(sample_rows):
    bulk = read_bulk_input(workbook_bytes(sample_rows))
    output, stats = generate_diagnostic_workbook(bulk, "ASIN (Informational only)", "B0PARENTA1")
    workbook = load_workbook(io.BytesIO(output), data_only=False, keep_links=False)

    assert workbook.sheetnames == [
        "广告组合预算", "Bulk", "BidingAdjustment", "广告位可视化", "搜索词模板",
        "词频分析(全部搜索词）", "词频分析 (不出单搜素词)", "否词库",
    ]
    assert stats.selected_values == ("B0PARENTA1",)
    assert stats.filtered_rows == 2
    assert stats.bidding_rows == 1
    assert stats.search_terms == 1
    assert workbook["Bulk"]["W15"].value == "B0PARENTA1"
    assert workbook["Bulk"]["W16"].value is None
    assert workbook["Bulk"]["L16"].value == "Parent-A Phrase"
    assert workbook["Bulk"].freeze_panes == "K15"
    assert workbook["Bulk"].auto_filter.ref == "A14:BE16"
    assert workbook["BidingAdjustment"].freeze_panes == "N6"
    assert workbook["BidingAdjustment"]["AH6"].value == "Placement Top"
    assert workbook["搜索词模板"]["A4"].value == "carry on luggage"
    assert workbook["搜索词模板"]["B4"].value == 1000
    assert workbook["搜索词模板"]["G4"].value == "=IFERROR(D4/C4,0)"
    assert workbook["词频分析(全部搜索词）"]["A2"].value in {"carry", "luggage", "on"}
    assert len(workbook["广告位可视化"]._charts) == 2


def test_multiple_parent_values_are_combined(sample_rows):
    bulk = read_bulk_input(workbook_bytes(sample_rows))
    output, stats = generate_diagnostic_workbook(
        bulk,
        "ASIN (Informational only)",
        ["B0PARENTA1", "B0OTHER001", "B0PARENTA1"],
    )
    workbook = load_workbook(io.BytesIO(output), data_only=False, keep_links=False)

    assert stats.selected_values == ("B0PARENTA1", "B0OTHER001")
    assert stats.filtered_rows == 3
    assert stats.search_terms == 2
    assert workbook["Bulk"]["W17"].value == "B0OTHER001"


def test_campaign_filter_uses_contains_match(sample_rows):
    bulk = read_bulk_input(workbook_bytes(sample_rows))
    _, stats = generate_diagnostic_workbook(
        bulk,
        "Campaign Name (Informational only)",
        "Parent-A",
    )
    assert stats.filtered_rows == 2


def test_missing_required_metric_is_rejected(sample_rows):
    headers = [header for header in HEADERS if header != "Spend"]
    spend_index = HEADERS.index("Spend")
    rows = [[value for index, value in enumerate(row) if index != spend_index] for row in sample_rows]
    with pytest.raises(InputValidationError, match="Spend"):
        read_bulk_input(workbook_bytes(rows, headers=headers))


def test_template_is_deidentified_and_available():
    assert Path(TEMPLATE_PATH).is_file()
    workbook = load_workbook(TEMPLATE_PATH, data_only=False, keep_links=False)
    assert workbook["Bulk"].max_row == 15
    assert workbook["Bulk"]["A15"].value is None
    assert workbook["NicheWord"]["A6"].value is None


def test_safe_filename_removes_invalid_characters():
    name = safe_filename('B0/ABC:*?"<>|')
    assert "/" not in name
    assert ":" not in name
    assert name.endswith(".xlsx")
