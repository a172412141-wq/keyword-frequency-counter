from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

from listing_optimizer import build_listing_result
from optimizer import optimize_title
from skill_engine import analyze_title, normalize_space


EXPECTED_COLUMNS = {
    "title": "Title",
    "brand": "Brand",
    "category": "Category",
}

TEMPLATE_HEADERS = ["Title", "Brand", "Category"]
LISTING_TEMPLATE_HEADERS = [
    "Title",
    "Brand",
    "Category",
    "Bullet1",
    "Bullet2",
    "Bullet3",
    "Bullet4",
    "Bullet5",
    "APlusContent",
]


def build_result(title: str, brand: str = "", category: str = "") -> dict[str, Any]:
    clean_title = normalize_space(title)
    clean_brand = normalize_space(brand)
    clean_category = normalize_space(category)
    issues = analyze_title(clean_title, clean_brand, clean_category)

    return {
        "original": clean_title,
        "status": "FAIL" if issues else "PASS",
        "issues": issues,
        "optimized_title": optimize_title(clean_title, clean_brand, clean_category),
    }


def read_title_rows(file_bytes: bytes) -> list[dict[str, str]]:
    try:
        dataframe = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise ValueError("Unable to read Excel file. Upload a valid .xlsx workbook.") from exc

    normalized_columns = {
        str(column).strip().lower(): column for column in dataframe.columns
    }
    if "title" not in normalized_columns:
        raise ValueError("Missing required column: Title")

    rows: list[dict[str, str]] = []
    for _, row in dataframe.iterrows():
        title = _cell_to_string(row.get(normalized_columns["title"]))
        if not title:
            continue

        rows.append(
            {
                "title": title,
                "brand": _optional_cell(row, normalized_columns, "brand"),
                "category": _optional_cell(row, normalized_columns, "category"),
            }
        )

    return rows


def read_listing_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        dataframe = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise ValueError("Unable to read Excel file. Upload a valid .xlsx workbook.") from exc

    normalized_columns = {
        str(column).strip().lower(): column for column in dataframe.columns
    }
    if "title" not in normalized_columns:
        raise ValueError("Missing required column: Title")

    rows: list[dict[str, Any]] = []
    for _, row in dataframe.iterrows():
        title = _cell_to_string(row.get(normalized_columns["title"]))
        if not title:
            continue

        bullets = [
            _optional_cell(row, normalized_columns, f"bullet{index}")
            for index in range(1, 6)
        ]
        rows.append(
            {
                "title": title,
                "brand": _optional_cell(row, normalized_columns, "brand"),
                "category": _optional_cell(row, normalized_columns, "category"),
                "bullets": [bullet for bullet in bullets if bullet],
                "aplus_content": _optional_cell(row, normalized_columns, "apluscontent"),
            }
        )

    return rows


def build_listing_batch_results(file_bytes: bytes) -> list[dict[str, Any]]:
    return [
        build_listing_result(
            row["title"],
            row.get("brand", ""),
            row.get("category", ""),
            row.get("bullets", []),
            row.get("aplus_content", ""),
        )
        for row in read_listing_rows(file_bytes)
    ]


def create_blank_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Title Template"
    sheet.append(TEMPLATE_HEADERS)

    header_fill = PatternFill("solid", fgColor="E0F2F1")
    header_font = Font(bold=True, color="0F172A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 28
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_listing_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listing Template"
    sheet.append(LISTING_TEMPLATE_HEADERS)

    header_fill = PatternFill("solid", fgColor="E0F2F1")
    header_font = Font(bold=True, color="0F172A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = [48, 22, 28, 42, 42, 42, 42, 42, 72]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _optional_cell(row: pd.Series, normalized_columns: dict[str, Any], name: str) -> str:
    column = normalized_columns.get(name)
    if column is None:
        return ""
    return _cell_to_string(row.get(column))


def _cell_to_string(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return normalize_space(str(value))
